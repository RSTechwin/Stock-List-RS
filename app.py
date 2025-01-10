from flask import Flask,send_file, render_template, request, redirect, url_for, session, send_from_directory,jsonify,flash
import os  
import pandas as pd  
import shutil
import tempfile
import time
import logging 
from io import BytesIO
import sqlite3
from openpyxl import load_workbook
import smtplib
import json
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pytz import timezone
from datetime import datetime
from math import ceil
from urllib.parse import quote, unquote
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'supersecretkey'

# Hardcoded credentials for login
USERNAME = "Admin"
PASSWORD = "@Dminrs123"

STAFFNAME="RsTechwin"
STAFFPASS="adminrs123"

# Define the path to the Excel file
SUPPLIER_FILE='supplier.json'
EXCEL_FILE_PATH = os.path.join("files", "stockList.xlsx")
LAST_EMAIL_FILE = os.path.join("files", "last_email_timestamp.json")


# Helper function to clean the data
def clean_data(df):
    # Replace NaN with '--'
    df = df.fillna('--')
    # Replace 'N/A' with '--'
    df = df.replace('N/A', '--')
    return df

@app.route('/restoreDashboard')
def restoreDashboard():
    if not os.path.exists(EXCEL_FILE_PATH):
        return "Stock data file not found."
    try:
        # Load and clean data
        df = clean_data(pd.read_excel(EXCEL_FILE_PATH))

        # Convert Stock Balance to numeric, with error handling
        df['Stock Balance'] = pd.to_numeric(df.get('Stock Balance', 0), errors='coerce').fillna(0).astype(int)

        # Handle Min. Qty: Convert to numeric, and leave blank if not present
        if 'Min. Qty' in df.columns:
            df['Min. Qty'] = pd.to_numeric(df['Min. Qty'], errors='coerce')
        else:
            df['Min. Qty'] = None

        # Debug: Print a sample of the Min. Qty column to verify correctness
        print("Debug - Min. Qty Sample:")
        print(df['Min. Qty'].head())

        # Filter for low-stock items (Stock Balance < Min. Qty) where Min. Qty is not blank
        low_stock = df[(df['Min. Qty'].notna()) & (df['Stock Balance'] < df['Min. Qty'])]

        # Paginate results
        page = int(request.args.get('page', 1))
        per_page = 30
        start_idx = (page - 1) * per_page
        paginated_data = low_stock.iloc[start_idx:start_idx + per_page]

        return render_template(
            'restoreDashboard.html',
            table_data=paginated_data.to_dict(orient='records'),
            page=page,
            total_pages=ceil(len(low_stock) / per_page)
        )
    except Exception as e:
        return f"Error processing the stock data: {e}"
    
def clean_data(df):
    # Implement your cleaning logic here
    return df

# Routes and Categories
def create_update_form(item_name, category_name):
    return f'''
        <form method="POST" action="{url_for('view_category_items', category_name=category_name)}" style="display:inline-block; margin:0;">
            <input type="hidden" name="item_name" value="{item_name}">
            <input type="number" name="stock_update" placeholder="Add/Deduct" required style="width:80px; margin-right:5px;">
            <button type="submit" style="padding:5px 10px; font-size:12px; background-color:#007bff; color:white; border:none; border-radius:4px; cursor:pointer;">Update</button>
        </form>
    '''

def load_suppliers():
    if not os.path.exists(SUPPLIER_FILE):
        return []
    with open(SUPPLIER_FILE, 'r') as file:
        return [s.strip() for s in json.load(file)]  # Normalize supplier names when loading

def save_suppliers(suppliers):
    # Normalize data before saving
    normalized_suppliers = sorted(set([s.strip() for s in suppliers]))  # Remove duplicates and sort alphabetically
    with open(SUPPLIER_FILE, 'w') as file:
        json.dump(normalized_suppliers, file, indent=4)
      
@app.route('/deleteStock', methods=['GET', 'POST'])
def deleteStock():
    if os.path.exists(EXCEL_FILE_PATH):
        # Load the stock list
        workbook = load_workbook(EXCEL_FILE_PATH)
        sheet1 = workbook["Sheet1"]
        update_sheet3 = workbook["updateSheet3"]

        # Convert Sheet1 to DataFrame for easier manipulation
        sheet1_data = list(sheet1.values)
        headers = sheet1_data[0]
        df = pd.DataFrame(sheet1_data[1:], columns=headers)

        # Replace NaN or empty cells with '--'
        df.fillna('--', inplace=True)

        # Get all item names
        item_names = df['Item Name'].tolist()

        if request.method == 'POST':
            # Get the item to delete
            item_to_delete = request.form['item_name']

            # Check if the item exists
            if item_to_delete in df['Item Name'].values:
                # Log the deletion in updateSheet3
                row_index = df[df['Item Name'] == item_to_delete].index[0]
                deleted_item_details = df.loc[row_index].to_dict()

                # Remove the item from the DataFrame
                df.drop(index=row_index, inplace=True)

                # Update Sheet1
                sheet1.delete_rows(row_index + 2)  # +2 for 1-indexing and header row

                # Log the deletion into updateSheet3
                new_row_index = update_sheet3.max_row + 1
                update_sheet3.cell(row=new_row_index, column=1, value=item_to_delete)
                update_sheet3.cell(row=new_row_index, column=2, value="Delete")
                update_sheet3.cell(row=new_row_index, column=3, value=0)  # Quantity
                update_sheet3.cell(row=new_row_index, column=4, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                update_sheet3.cell(row=new_row_index, column=5, value=deleted_item_details.get('Stock Balance', 0))

                # Save the workbook
                workbook.save(EXCEL_FILE_PATH)
                push_to_github() 

                # Redirect back with success message
                return redirect(url_for('deleteStock', message=f"Item '{item_to_delete}' deleted successfully."))
            else:
                # Redirect back with error message
                return redirect(url_for('deleteStock', message="Item not found."))

        # Render the deleteStock.html template
        return render_template('deleteStock.html', item_names=item_names)
    else:
        return "Excel file not found."

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username == USERNAME and password == PASSWORD:
            session['user'] = username
            session['role'] = 'admin'
            return redirect(url_for('dashboard'))
        elif username == STAFFNAME and password == STAFFPASS:
            session['user'] = username
            session['role'] = 'staff'
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error="Invalid credentials!")
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'user' in session:  # Check if the user is logged in
        return render_template('stockDashboard.html')  # Show the dashboard
    else:
        return redirect(url_for('login'))  # Redirect to login if not logged in
    
def log_stock_update(item_name, action, quantity, date, who_taken=None, site_name=None, remarks=None):
    try:
        # Load stock file to get the current stock BEFORE update
        stock_df = pd.DataFrame(columns=['Item Name', 'Stock Balance'])
        if os.path.exists(EXCEL_FILE_PATH):
            stock_df = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Sheet1')  # Specify sheet for clarity
            stock_entry = stock_df[stock_df['Item Name'].str.strip().str.lower() == item_name.strip().lower()]
            current_stock_before_update = stock_entry.iloc[0]['Stock Balance'] if not stock_entry.empty else 0
        else:
            current_stock_before_update = 0  # Default to 0 if stock file doesn't exist

        # Load the log file or create a new one if it doesn't exist
        log_df = pd.DataFrame(columns=['Item Name', 'Action', 'Quantity', 'Date', 'Who Taken', 'Site Name', 'Remarks', 'Current Stock Before Update'])
        if os.path.exists("sheet1"):
            log_df = pd.read_excel("sheet1")

        # Add the new log entry
        new_log = {
            'Item Name': item_name,
            'Action': action,
            'Quantity': quantity if action != "Delete" else 0,
            'Date': date,
            'Who Taken': who_taken,
            'Site Name': site_name,
            'Remarks': remarks,
            'Current Stock Before Update': current_stock_before_update,  # Log stock BEFORE the update
        }
        log_df = pd.concat([log_df, pd.DataFrame([new_log])], ignore_index=True)

        # Save the updated log file
        with pd.ExcelWriter("sheet1", engine='openpyxl', mode='w') as writer:
            log_df.to_excel(writer, index=False)

        print(f"Logged update for '{item_name}' - Action: {action}, Quantity: {quantity}, "
              f"Who Taken: {who_taken}, Site Name: {site_name}, Remarks: {remarks}, "
              f"Current Stock Before Update: {current_stock_before_update}")

    except Exception as e:
        print(f"Error in log_stock_update: {e}")

@app.route('/get_stock_updates', methods=['GET'])
def get_stock_updates():
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            # Load the workbook and relevant sheets
            workbook = load_workbook(EXCEL_FILE_PATH, data_only=True)
            sheet1 = workbook["Sheet1"]  # Contains current stock
            sheet2 = workbook["updateSheet3"]  # Contains transaction logs

            # Convert Sheet1 and Sheet2 to DataFrames
            sheet1_data = list(sheet1.values)
            sheet1_columns = sheet1_data.pop(0)  # Extract headers
            stock_df = pd.DataFrame(sheet1_data, columns=sheet1_columns)

            sheet2_data = list(sheet2.values)
            sheet2_columns = sheet2_data.pop(0)  # Extract headers
            updates_df = pd.DataFrame(sheet2_data, columns=sheet2_columns)

            # Normalize column names for consistency
            stock_df.columns = stock_df.columns.str.strip()
            updates_df.columns = updates_df.columns.str.strip()

            # Ensure required columns exist in both sheets
            if "Item Name" not in updates_df.columns or "Item Name" not in stock_df.columns:
                return jsonify({'error': 'Missing required columns in the Excel file'}), 500

            # Merge current stock (from Sheet1) into the updates (from updateSheet3)
            updates_df = updates_df.merge(
                stock_df[['Item Name', 'Stock Balance']],
                how='left',
                on='Item Name'
            )

            # Add minus sign for "Out" actions and ensure numeric Quantity
            updates_df['Quantity'] = pd.to_numeric(updates_df['Quantity'], errors='coerce').fillna(0).astype(int)
            updates_df.loc[updates_df['Action'].str.lower() == 'out', 'Quantity'] *= -1

            # Select and rename the necessary columns
            updates_df = updates_df[[
                'Item Name', 'Action', 'Quantity', 'Date', 'Site Name', 'Who Taken', 'Stock Balance'
            ]].rename(columns={
                'Item Name': 'Item Name',
                'Action': 'Action',
                'Quantity': 'Stock Out',
                'Date': 'Date',
                'Site Name': 'Site Name',
                'Who Taken': 'Taken By',
                'Stock Balance': 'Current Stock'
            })

            # Replace NaN values with defaults for display
            updates_df.fillna({
                'Item Name': '--',
                'Action': '--',
                'Stock Out': 0,
                'Date': '--',
                'Site Name': '--',
                'Taken By': '--',
                'Current Stock': 0
            }, inplace=True)

            # Convert to JSON format for frontend
            log_data = updates_df.to_dict(orient='records')
            return jsonify(log_data), 200

        return jsonify([]), 200
    except Exception as e:
        print(f"Error in get_stock_updates: {e}")
        return jsonify({'error': 'Server error while fetching stock updates'}), 500

@app.route('/get_out_stock_summary', methods=['GET'])
def get_out_stock_summary():
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            # Load the Excel workbook
            workbook = load_workbook(EXCEL_FILE_PATH, data_only=True)
            transaction_sheet = workbook["Transaction"]  # Transaction log
            stock_sheet = workbook["Sheet1"]  # Current stock data

            # Convert sheets to DataFrames
            transaction_data = list(transaction_sheet.values)
            transaction_columns = transaction_data.pop(0)  # Extract column headers
            transaction_df = pd.DataFrame(transaction_data, columns=transaction_columns)

            stock_data = list(stock_sheet.values)
            stock_columns = stock_data.pop(0)  # Extract column headers
            stock_df = pd.DataFrame(stock_data, columns=stock_columns)

            # Normalize column names
            transaction_df.columns = transaction_df.columns.str.strip()
            stock_df.columns = stock_df.columns.str.strip()

            # Replace missing values with '--'
            transaction_df.fillna({'Site Name': '--', 'Who Taken': '--', 'Remark': '--'}, inplace=True)

            # Ensure numeric values for stock balance and quantity
            transaction_df['Quantity'] = pd.to_numeric(transaction_df['Quantity'], errors='coerce').fillna(0).astype(int)
            stock_df['Stock Balance'] = pd.to_numeric(stock_df.get('Stock Balance', 0), errors='coerce').fillna(0).astype(int)

            # Merge current stock data from Sheet1 into the Transaction sheet
            merged_df = transaction_df.merge(
                stock_df[['Item Name', 'Stock Balance']],
                how='left',
                on='Item Name'
            )

            # Add a minus sign for "Out" actions
            merged_df.loc[merged_df['Action'].str.lower() == 'out', 'Quantity'] *= -1

            # Rename columns for better readability
            merged_df = merged_df.rename(columns={
                'Item Name': 'Item Name',
                'Action': 'Action',
                'Date': 'Date',
                'Quantity': 'Stock Change',
                'Site Name': 'Site Name',
                'Who Taken': 'Taken By',
                'Remark': 'Remarks',
                'Stock Balance': 'Current Stock'
            })

            # Replace any remaining NaN values with default values
            merged_df.fillna({'Site Name': '--', 'Taken By': '--', 'Remarks': '--', 'Current Stock': 0}, inplace=True)

            # Convert to JSON for the frontend
            response_data = merged_df.to_dict(orient='records')
            return jsonify(response_data), 200

        return jsonify({'error': 'Excel file not found'}), 404
    except Exception as e:
        print(f"Error in get_out_stock_summary: {e}")
        return jsonify({'error': 'Server error while fetching out stock summary'}), 500

@app.route('/get_category_items', methods=['GET', 'POST'])
def get_category_items():
    """
    Fetch and update category items. Updates Sheet1 and updateSheet3 in the Excel file.
    """
    if not os.path.exists(EXCEL_FILE_PATH):
        return "Stock data file not found.", 404

    # Load the Excel file
    workbook = load_workbook(EXCEL_FILE_PATH)
    sheet1 = workbook["Sheet1"]
    update_sheet3 = workbook["updateSheet3"]

    # Convert Sheet1 to a DataFrame
    sheet1_data = list(sheet1.values)
    headers = sheet1_data[0]
    df = pd.DataFrame(sheet1_data[1:], columns=headers)

    # Clean data by replacing NaN and empty cells with '--'
    df = df.fillna('--')

    if request.method == 'POST':
        try:
            # Extract data from the form
            item_name = request.form['item_name'].strip()
            quantity = int(request.form['quantity'])
            action = request.form['action']  # "Add" or "Deduct"
            remarks = request.form.get('remarks', '').strip()

            # Check if item exists in Sheet1
            item_mask = df['Item Name'].str.strip().str.lower() == item_name.lower()
            if item_mask.any():
                # Update existing item
                row_idx = item_mask.idxmax() + 2  # Excel rows are 1-indexed and include headers
                current_stock = int(sheet1.cell(row=row_idx, column=headers.index('Stock Balance') + 1).value or 0)

                if action.lower() == 'add':
                    new_stock = current_stock + quantity
                elif action.lower() == 'deduct':
                    new_stock = max(0, current_stock - quantity)
                else:
                    return "Invalid action.", 400

                # Update Sheet1
                sheet1.cell(row=row_idx, column=headers.index('Stock Balance') + 1, value=new_stock)
                sheet1.cell(row=row_idx, column=headers.index('Last Updated') + 1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

                # Log the update in updateSheet3
                new_row_idx = update_sheet3.max_row + 1
                update_sheet3.cell(new_row_idx, 1, value=item_name)
                update_sheet3.cell(new_row_idx, 2, value=action)
                update_sheet3.cell(new_row_idx, 3, value=quantity)
                update_sheet3.cell(new_row_idx, 4, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                update_sheet3.cell(new_row_idx, 5, value=new_stock)
                update_sheet3.cell(new_row_idx, 6, value=request.form.get('who_taken', '--'))
                update_sheet3.cell(new_row_idx, 7, value=request.form.get('site_name', '--'))
                update_sheet3.cell(new_row_idx, 8, value=remarks)

                # Save changes
                workbook.save(EXCEL_FILE_PATH)

                # Push to GitHub only after saving changes
                push_to_github()

                return jsonify({"status": "success", "message": f"Item '{item_name}' updated successfully.", "new_stock": new_stock})
            else:
                return jsonify({"status": "error", "message": "Item not found."}), 404
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    # For GET requests, return all items
    items = df.fillna('--').to_dict(orient='records')
    return jsonify(items)



app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'static/images')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
    is_allowed = '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    print(f"Checking file: {filename}, Allowed: {is_allowed}")
    return is_allowed

import re

def custom_secure_filename(filename):
    """
    Custom function to sanitize filenames while allowing specific characters like [].
    """
    # Allow alphanumeric, underscores, hyphens, dots, and brackets
    filename = re.sub(r'[^\w\.\-()\[\]]+', '', filename)
    return filename

@app.route('/upload_image', methods=['POST'])
def upload_image():
    print("Request form data:", request.form)
    print("Request files data:", request.files)

    app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'static/images')

    if 'file' not in request.files:
        print("No file part in the request.")
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        print("No file selected.")
        return jsonify({'error': 'No selected file'}), 400

    if file and allowed_file(file.filename):
        original_filename = file.filename
        sanitized_filename = custom_secure_filename(original_filename.replace(" ", "_"))
        
        item_name = request.form.get('item_name')
        if not item_name:
            print("Item name not provided.")
            return jsonify({'error': 'Item name not provided'}), 400

        save_path = os.path.join(app.config['UPLOAD_FOLDER'], sanitized_filename)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)  # Ensure directory exists
        file.save(save_path)  # Save the file
        print(f"File saved at: {save_path}")

        return jsonify({'success': True, 'path': url_for('static', filename=f'images/{sanitized_filename}')})
    else:
        print("File type not allowed.")
        return jsonify({'error': 'File not allowed'}), 400

import re

def sanitize_filename(filename):
    """
    Sanitize the filename by replacing invalid characters with underscores
    or other acceptable characters.
    """
    # Replace invalid characters with underscores or acceptable replacements
    sanitized_name = re.sub(r'[<>:"/\\|?*]', '', filename)  # Replace invalid characters with ''
    return sanitized_name.strip()

def get_image_path(item_name):
    """
    Matches item names with sanitized image file names and returns the relative path to the image.
    """
    image_folder = os.path.join('static', 'images')
    # Sanitize the item name to match the sanitized filenames
    sanitized_item_name = sanitize_filename(item_name).replace(' ', '_').lower()
    for ext in ['.jpg', '.jpeg', '.png', '.gif']:
        image_path = os.path.join(image_folder, f"{sanitized_item_name}{ext}")
        if os.path.exists(image_path):
            return url_for('static', filename=f"images/{sanitized_item_name}{ext}")
    return None  # No image found

def clean_data(df):
    df = df.fillna('--')  # Replace NaN with '--'
    df = df.replace('N/A', '--')  # Replace 'N/A' with '--'
    return df

@app.route('/view-categories', methods=['GET', 'POST'])
def view_categories():
    """List all sub-categories and allow users to choose one or multiple to view."""
    if os.path.exists(EXCEL_FILE_PATH):
        try:
            # Load the Excel file
            df = pd.read_excel(EXCEL_FILE_PATH)
            df = clean_data(df)

            # Count items in each sub-category
            sub_category_counts = df['Sub Category'].value_counts().to_dict()

            if request.method == 'POST':
                # Get selected sub-categories from the form
                selected_sub_categories = request.form.getlist('sub_categories')
                if selected_sub_categories:
                    # Redirect to the view_category_items route with selected sub-categories
                    selected_sub_categories_encoded = ",".join(selected_sub_categories)
                    return redirect(url_for('view_category_items', sub_category_names=selected_sub_categories_encoded))

            return render_template(
                'view_categories.html',
                sub_categories=sub_category_counts
            )
        except Exception as e:
            return f"Error loading categories: {e}"
    return "Excel file not found."

@app.route('/view-category-items', methods=['GET', 'POST'])
def view_category_items():
    """Display items for one or multiple sub-categories, including handling stock updates and logging transactions."""
    sub_category_names = request.args.get('sub_category_names', '')  # Get selected sub-categories
    selected_sub_categories = sub_category_names.split(',') if sub_category_names else []

    if os.path.exists(EXCEL_FILE_PATH):
        try:
            # Load the Excel file
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet1 = workbook["Sheet1"]  # Main stock data
            update_sheet3 = workbook["updateSheet3"]  # Update log
            transaction_sheet = workbook["Transaction"]  # Transaction log

            # Convert the sheet to a DataFrame
            sheet1_data = list(sheet1.values)
            headers = sheet1_data[0]
            df = pd.DataFrame(sheet1_data[1:], columns=headers)

            # Ensure required columns exist in the DataFrame
            required_columns = ['Sub Category', 'Item Name', 'Stock Balance', 'Supplier']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = None  # Add missing columns with default values

            if request.method == 'POST':
                item_name = request.form.get('item_name')
                stock_update = request.form.get('stock_update')
                who_taken = request.form.get('who_taken', '--')
                site_name = request.form.get('site_name', '--')
                remarks = request.form.get('remarks', '--')

                if not item_name or not stock_update:
                    return "Invalid form data."

                try:
                    stock_update = int(stock_update)
                except ValueError:
                    return "Invalid stock update value."

                # Determine the action based on the stock_update value
                action = 'Add' if stock_update > 0 else 'Deduct'

                # Find the item and update its stock balance
                mask = df['Item Name'].str.strip().str.lower() == item_name.strip().lower()
                if not df.loc[mask, 'Stock Balance'].empty:
                    current_stock = pd.to_numeric(df.loc[mask, 'Stock Balance'].values[0], errors='coerce') or 0
                    new_stock_balance = max(0, current_stock + stock_update)
                    df.loc[mask, 'Stock Balance'] = new_stock_balance

                    # Update the Excel sheet (Sheet1)
                    for row_index, row in df.iterrows():
                        for col_index, value in enumerate(row):
                            sheet1.cell(row=row_index + 2, column=col_index + 1, value=value)

                    # Log the transaction in updateSheet3
                    new_row_index = update_sheet3.max_row + 1
                    update_sheet3.cell(new_row_index, 1, value=item_name)
                    update_sheet3.cell(new_row_index, 2, value=action)
                    update_sheet3.cell(new_row_index, 3, value=stock_update)  # Quantity (Add/Deduct)
                    update_sheet3.cell(new_row_index, 4, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))  # Current date
                    update_sheet3.cell(new_row_index, 5, value=new_stock_balance)  # Current stock
                    update_sheet3.cell(new_row_index, 6, value=who_taken)
                    update_sheet3.cell(new_row_index, 7, value=site_name)
                    update_sheet3.cell(new_row_index, 8, value=remarks)

                    # Log the transaction in the Transaction sheet
                    transaction_row_index = transaction_sheet.max_row + 1
                    transaction_sheet.cell(transaction_row_index, 1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    transaction_sheet.cell(transaction_row_index, 2, value=item_name)  # Item Name
                    transaction_sheet.cell(transaction_row_index, 3, value=action)  # Action
                    transaction_sheet.cell(transaction_row_index, 6, value=stock_update)  # Correct Quantity
                    transaction_sheet.cell(transaction_row_index, 4, value=site_name)  # Site Name (if applicable)
                    transaction_sheet.cell(transaction_row_index, 7, value=who_taken)  # Who Taken
                    transaction_sheet.cell(transaction_row_index, 5, value=remarks)  # Remarks
                    transaction_sheet.cell(transaction_row_index, 8, value=new_stock_balance)  # Current stock

                    # Save the workbook
                    workbook.save(EXCEL_FILE_PATH)

                    # Redirect back to the same filtered view with the selected subcategories
                    return redirect(
                        url_for(
                            'view_category_items',
                            sub_category_names=request.form.get('sub_category_names', ''),
                            search=request.args.get('search', ''),
                            supplier=request.args.get('supplier', ''),
                            stock_status=request.args.get('stock_status', '')
                        )
                    )

            # Filter items based on selected sub-categories
            if selected_sub_categories:
                filtered_df = df[df['Sub Category'].isin(selected_sub_categories)]
            else:
                filtered_df = df

            filtered_df['Picture'] = filtered_df['Item Name'].apply(lambda x: get_image_path(x))

            # Handle additional filters for search, supplier, and stock status
            search_query = request.args.get('search', '').strip().lower()
            stock_status = request.args.get('stock_status', '').strip()
            supplier_filter = request.args.get('supplier', '').strip().lower()

            if search_query:
                filtered_df = filtered_df[filtered_df['Item Name'].str.lower().str.contains(search_query)]

            filtered_df['Stock Balance'] = pd.to_numeric(filtered_df['Stock Balance'], errors='coerce').fillna(0).astype(int)

            if stock_status == 'Stock Low':
                filtered_df = filtered_df[filtered_df['Stock Balance'] <= 10]
            elif stock_status == 'Stock OK':
                filtered_df = filtered_df[filtered_df['Stock Balance'] > 10]

            if supplier_filter:
                filtered_df = filtered_df[filtered_df['Supplier'].str.lower() == supplier_filter]

            # Pagination logic
            page = request.args.get('page', 1, type=int)
            per_page = 10
            start_idx = (page - 1) * per_page
            paginated_data = filtered_df.iloc[start_idx:start_idx + per_page][[
                'Sub Category',
                'Item Name',
                'Supplier',
                'Stock Balance',
                'Unit',
                'Local Cost(RM)',
                'Average Purchase Price',
                'Picture'
            ]]

            total_pages = ceil(len(filtered_df) / per_page)

            # Get unique, valid suppliers dynamically for the dropdown
            unique_suppliers = (
                filtered_df['Supplier']
                .dropna()
                .apply(lambda x: x.strip())
                .unique()
            )
            unique_suppliers = [supplier for supplier in unique_suppliers if supplier and supplier != '--']
            
            # Render the template
            return render_template(
                'view_category_items.html',
                sub_category_name=", ".join(selected_sub_categories),
                table_data=paginated_data.to_dict(orient='records'),
                page=page,
                total_pages=total_pages,
                unique_suppliers=unique_suppliers,
                selected_supplier=supplier_filter
            )
        except Exception as e:
            return f"Error displaying sub-category items: {e}"
    return "Excel file not found."

@app.route('/download-categories', methods=['GET'])
def download_categories_excel():
    """Download filtered Excel file for multiple sub-categories."""
    try:
        sub_category_names = request.args.get('sub_category_names', '')  # Get selected sub-categories
        selected_sub_categories = sub_category_names.split(',') if sub_category_names else []

        if not os.path.exists(EXCEL_FILE_PATH):
            return "Stock data file not found.", 404

        # Load and clean data
        df = pd.read_excel(EXCEL_FILE_PATH)
        df = clean_data(df)

        df['Stock Balance'] = pd.to_numeric(df['Stock Balance'], errors='coerce').fillna(0).astype(int)
        # Filter by multiple sub-categories
        if selected_sub_categories:
            filtered_df = df[df['Sub Category'].isin(selected_sub_categories)]
        else:
            filtered_df = df

        # Apply additional filters
        search_query = request.args.get('search', '').strip().lower()
        if search_query:
            filtered_df = filtered_df[filtered_df['Item Name'].str.lower().str.contains(search_query)]

        stock_status = request.args.get('stock_status', '').strip()
        if stock_status == 'Stock Low':
            filtered_df = filtered_df[filtered_df['Stock Balance'] <= 10]
        elif stock_status == 'Stock OK':
            filtered_df = filtered_df[filtered_df['Stock Balance'] > 10]

        supplier_filter = request.args.get('supplier', '').strip().lower()
        if supplier_filter:
            filtered_df = filtered_df[filtered_df['Supplier'].str.lower() == supplier_filter]

        # Save the filtered data to a temporary file
        temp_dir = tempfile.mkdtemp()
        temp_file = os.path.join(temp_dir, "filtered_categories.xlsx")
        filtered_df.to_excel(temp_file, index=False)

        return send_file(temp_file, as_attachment=True, download_name="filtered_categories.xlsx")
    except Exception as e:
        return f"Error: {e}", 500

from datetime import datetime
from pytz import timezone

@app.route('/suppliers', methods=['GET'])
def get_suppliers():
    suppliers = load_suppliers()
    return jsonify(suppliers)

@app.route('/suppliers', methods=['POST'])
def add_supplier():
    new_supplier = request.json.get('supplier_name', '').strip()
    if new_supplier:
        suppliers = load_suppliers()
        if new_supplier not in suppliers:
            suppliers.append(new_supplier)
            save_suppliers(suppliers)
            push_to_github() 
            return jsonify({'status': 'success', 'message': f"Supplier '{new_supplier}' added.", 'suppliers': suppliers})
    return jsonify({'status': 'error', 'message': 'Invalid or duplicate supplier.'})

@app.route('/suppliers', methods=['DELETE'])
def delete_supplier():
    supplier_to_delete = request.json.get('supplier_name', '').strip()
    suppliers = load_suppliers()

    # Normalize supplier names
    normalized_suppliers = [s.strip().lower() for s in suppliers]
    normalized_supplier_to_delete = supplier_to_delete.lower()

    if normalized_supplier_to_delete in normalized_suppliers:
        index_to_remove = normalized_suppliers.index(normalized_supplier_to_delete)
        removed_supplier = suppliers.pop(index_to_remove)
        save_suppliers(suppliers)
        push_to_github() 
        return jsonify({'status': 'success', 'message': f"Supplier '{removed_supplier}' deleted.", 'suppliers': suppliers})

    return jsonify({'status': 'error', 'message': 'Supplier not found.'})

@app.route('/enter-out-stock', methods=['GET', 'POST'])
def enterOutStock():
    success_message = ""

    if os.path.exists(EXCEL_FILE_PATH):
        try:
            # Load the Excel file
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet1 = workbook['Sheet1']  # Current stock list
            update_sheet3 = workbook['updateSheet3']  # Dashboard sheet

            # Extract headers from Sheet1
            headers1 = {sheet1.cell(row=1, column=i + 1).value: i + 1 for i in range(sheet1.max_column)}

            # Initialize required columns if they don't exist
            for col in ['Sub Category', 'Item Name', 'Stock Balance', 'Last Updated', 'Supplier', 'Unit']:
                if col not in headers1:
                    headers1[col] = sheet1.max_column + 1
                    sheet1.cell(row=1, column=headers1[col]).value = col

            # Prepare a list of sub-categories for the dropdown
            df = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Sheet1')
            if 'Sub Category' in df.columns:
                sub_categories = sorted(df['Sub Category'].dropna().unique().tolist())
            else:
                sub_categories = []
        except Exception as e:
            return f"Error loading Excel file: {e}"
    else:
        return "Excel file not found."

    if request.method == 'POST':
        # Get form data
        item_name = request.form['item_name'].strip()
        sub_category = request.form['sub_category'].strip()
        quantity = int(request.form['quantity'])
        supplier = request.form.get('supplier', '').strip()
        unit = request.form.get('unit', '').strip()  # New unit field
        who_take = request.form.get('who_take', '').strip()
        site_name = request.form.get('site_name', '').strip()
        remarks = request.form.get('remarks', '').strip()

        # Assign the current date
        current_date = datetime.now()
        formatted_date = current_date.astimezone(timezone("GMT")).strftime("%a, %d %b %Y %H:%M:%S %Z")

        # Check if item exists in Sheet1
        item_found = False
        for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=headers1['Item Name'], max_col=headers1['Item Name']):
            cell = row[0]
            if cell.value and cell.value.strip().lower() == item_name.lower():
                item_found = True
                current_row = cell.row
                current_quantity = sheet1.cell(row=current_row, column=headers1['Stock Balance']).value or 0

                # Update stock
                new_stock_balance = current_quantity + quantity
                sheet1.cell(row=current_row, column=headers1['Stock Balance']).value = new_stock_balance
                sheet1.cell(row=current_row, column=headers1['Last Updated']).value = formatted_date
                sheet1.cell(row=current_row, column=headers1['Supplier']).value = supplier
                sheet1.cell(row=current_row, column=headers1['Unit']).value = unit  # Update unit field
                success_message = f"Stock for '{item_name}' has been successfully updated."
                break

        # If item is not found, add it as a new item
        if not item_found:
            new_row = sheet1.max_row + 1
            sheet1.cell(row=new_row, column=headers1['Sub Category']).value = sub_category
            sheet1.cell(row=new_row, column=headers1['Item Name']).value = item_name
            sheet1.cell(row=new_row, column=headers1['Stock Balance']).value = quantity
            sheet1.cell(row=new_row, column=headers1['Last Updated']).value = formatted_date
            sheet1.cell(row=new_row, column=headers1['Supplier']).value = supplier
            sheet1.cell(row=new_row, column=headers1['Unit']).value = unit  # Add unit field
            success_message = f"New stock item '{item_name}' has been successfully added."

        # Update updateSheet3
        update_headers3 = {update_sheet3.cell(row=1, column=i + 1).value: i + 1 for i in range(update_sheet3.max_column)}

        # Add row to updateSheet3
        new_row_update = update_sheet3.max_row + 1
        # Safely populate each column if it exists
        if 'Item Name' in update_headers3:
            update_sheet3.cell(row=new_row_update, column=update_headers3['Item Name']).value = item_name
        if 'Action' in update_headers3:
            update_sheet3.cell(row=new_row_update, column=update_headers3['Action']).value = "Add" if quantity > 0 else "Deduct"
        if 'Quantity' in update_headers3:
            update_sheet3.cell(row=new_row_update, column=update_headers3['Quantity']).value = quantity
        if 'Date' in update_headers3:
            update_sheet3.cell(row=new_row_update, column=update_headers3['Date']).value = formatted_date
        if 'Current Stock' in update_headers3:
            update_sheet3.cell(row=new_row_update, column=update_headers3['Current Stock']).value = new_stock_balance if item_found else quantity
        if 'Who Take' in update_headers3:
            update_sheet3.cell(row=new_row_update, column=update_headers3['Who Take']).value = who_take
        if 'Site Name' in update_headers3:
            update_sheet3.cell(row=new_row_update, column=update_headers3['Site Name']).value = site_name
        if 'Remarks' in update_headers3:
            update_sheet3.cell(row=new_row_update, column=update_headers3['Remarks']).value = remarks

        # Save the updated Excel file
        try:
            workbook.save(EXCEL_FILE_PATH)
            push_to_github() 
        except Exception as e:
            return f"Error saving Excel file: {e}"

        # Log the update
        log_stock_update(
            item_name=item_name,
            action="Add" if quantity > 0 else "Deduct",
            quantity=quantity,
            date=formatted_date
        )

        return redirect(url_for('enterOutStock', message=success_message))

    return render_template('enterOutStock.html', sub_categories=sub_categories)

def should_send_email():
    """
    Check if an email should be sent based on the last sent timestamp
    """
    if not os.path.exists(LAST_EMAIL_FILE):
        return True
    
    try:
        with open(LAST_EMAIL_FILE, 'r') as f:
            last_email_data = json.load(f)
        
        last_email_time = datetime.fromisoformat(last_email_data.get('last_email_timestamp', '1900-01-01'))
        
        return datetime.now() - last_email_time > timedelta(days=0.5)
    
    except (FileNotFoundError, json.JSONDecodeError, ValueError):
        return True

def update_last_email_timestamp():
    """
    Update the timestamp of the last sent email
    """
    os.makedirs(os.path.dirname(LAST_EMAIL_FILE), exist_ok=True)
    
    with open(LAST_EMAIL_FILE, 'w') as f:
        json.dump({
            'last_email_timestamp': datetime.now().isoformat()
        }, f)

SENDER_EMAIL = "hockweipang@gmail.com"
RECEIVER_EMAIL = "hockweip@gmail.com"
Email_PASSWORD = "nxvt jgnw xics lnxr" 

@app.route('/out_stock_summary', methods=['GET', 'POST'])
def out_stock_summary():
    page = request.args.get('page', type=int, default=1)
    items_per_page = 30
    date_filter = request.args.get('date_filter', '')
    site_filter = request.args.get('site_filter', '').strip().lower()
    action_filter = request.args.get('action_filter', '').strip().lower()  # Normalize to lowercase
    item_filter = request.args.get('item_filter', '').strip().lower()  # Normalize to lowercase

    if os.path.exists(EXCEL_FILE_PATH):
        try:
            workbook = load_workbook(EXCEL_FILE_PATH, data_only=True)
            sheet1 = workbook["Sheet1"]
            sheet2 = workbook["Transaction"]

            sheet1_data = sheet1.values
            sheet1_columns = next(sheet1_data)
            stock_df = pd.DataFrame(sheet1_data, columns=sheet1_columns)

            sheet2_data = sheet2.values
            sheet2_columns = next(sheet2_data)
            transaction_df = pd.DataFrame(sheet2_data, columns=sheet2_columns)

            # Remove duplicate rows in the DataFrame
            transaction_df.drop_duplicates(inplace=True)

            # Normalize Action column to lowercase
            transaction_df['Action'] = transaction_df['Action'].str.strip().str.lower()
            transaction_df['Action'] = transaction_df['Action'].replace({
                'in': 'add',
                'out': 'deduct'
            })

            # Handle Date Parsing
            transaction_df['Date'] = transaction_df['Date'].apply(
                lambda x: f"{x} 00:00:00" if isinstance(x, str) and len(x) == 10 else x
            )
            transaction_df['Date'] = pd.to_datetime(transaction_df['Date'], errors='coerce')
            transaction_df['Formatted Date'] = transaction_df['Date'].dt.strftime('%Y-%m-%d %H:%M:%S').fillna('--')

            # Apply Filters
            if date_filter:
                specific_date = pd.to_datetime(date_filter, errors='coerce')
                transaction_df = transaction_df[transaction_df['Date'].dt.date == specific_date.date()]

            if site_filter:
                transaction_df = transaction_df[transaction_df['Site Name'].str.lower().str.contains(site_filter, na=False)]

            if action_filter:
                transaction_df = transaction_df[transaction_df['Action'] == action_filter]

            if item_filter:
                transaction_df = transaction_df[transaction_df['Item Name'].str.lower().str.contains(item_filter, na=False)]

            # Merge with Stock Data
            stock_column_name = "Stock Balance"
            transaction_df = transaction_df.merge(
                stock_df[['Item Name', stock_column_name]],
                how='left',
                on='Item Name'
            )

            # Adjust Quantity for Deduction
            transaction_df.loc[transaction_df['Action'] == 'deduct', 'Quantity'] = (
                transaction_df.loc[transaction_df['Action'] == 'deduct', 'Quantity'] * -1
            )

            # Select Relevant Columns
            transaction_df = transaction_df[['Item Name', 'Action', 'Formatted Date', 'Quantity', 'Site Name', 'Who Taken', 'Remark', stock_column_name]]
            transaction_df = transaction_df.rename(columns={
                'Item Name': 'Item Name',
                'Action': 'Action',
                'Formatted Date': 'Formatted Date',
                'Quantity': 'Quantity',
                'Site Name': 'Site Name',
                'Who Taken': 'Taken By',
                'Remark': 'Remark',
                stock_column_name: 'Current Stock'
            })

            # Replace None or NaN values with '--'
            transaction_df = transaction_df.fillna('--')

            # Pagination Logic
            total_items = len(transaction_df)
            total_pages = (total_items + items_per_page - 1) // items_per_page
            start_idx = (page - 1) * items_per_page
            end_idx = min(start_idx + items_per_page, total_items)

            paginated_data = transaction_df.iloc[start_idx:end_idx]
            table_html = paginated_data.to_html(
                classes='table table-striped', index=False, escape=False
            )

            # Load Sites for Filters
            site_file = 'site.json'
            sites = load_json_data(site_file) if os.path.exists(site_file) else []

            return render_template(
                'out_stock_summary.html',
                table_data=table_html,
                page=page,
                total_pages=total_pages,
                date_filter=date_filter,
                site_filter=site_filter,
                action_filter=action_filter,
                item_filter=item_filter,
                sites=sites
            )

        except Exception as e:
            return f"Error processing the stock update log: {e}"

    return "The stock file was not found."

def send_email(table_html):
    try:
        # Create email message
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = RECEIVER_EMAIL
        msg['Subject'] = "Stock Alert: Items Below Threshold"

        # Add HTML content to the email
        body = f"""
        <html>
             <body>
                <h2 style="color: #2e6c80;">Stock Alert: The following items are below the threshold</h2>
                {table_html}
                <p style="color: grey; font-size: 12px;">This is an automated message. Please do not reply.</p>
            </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))

        # Send the email via SMTP server
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SENDER_EMAIL, Email_PASSWORD)
        server.send_message(msg)
        server.quit()

        print("Email sent successfully!")
    except Exception as e:
        print(f"Failed to send email: {e}")

@app.route('/download-sheet1')
def download_sheet1():
    try:
        # Ensure the file exists
        if os.path.exists(EXCEL_FILE_PATH):
            # Load the Excel file into a DataFrame
            sheet1 = pd.read_excel(EXCEL_FILE_PATH, sheet_name="Sheet1")
            
            # List of columns to drop
            columns_to_drop = [
                "Unnamed: 7", "Local Cost(RM)", "Local Item Link", "China Cost (楼)",
                "Price Converted", "Min. Qty", "Unnamed: 15", "Unnamed: 16",
                "History Total IN", "History Total OUT", "History Shift To",
                "History Return", "Last Updated", "Stock Out", "Packing/Size", "Qty","Average Purchase Price"
            ]
            
            # Drop the specified columns if they exist
            sheet1 = sheet1.drop(columns=[col for col in columns_to_drop if col in sheet1.columns], errors="ignore")
            
            # Save the modified DataFrame to a new Excel file
            modified_file_path = "files/newStockList.xlsx"
            sheet1.to_excel(modified_file_path, index=False)
            
            # Send the modified file to the user
            return send_file(modified_file_path, as_attachment=True)
        else:
            return "File not found!", 404
    except Exception as e:
        return f"Error: {e}", 500

@app.route('/logout')
def logout():
    session.pop('user', None)  # Clear the session
    return redirect(url_for('login'))  # Redirect to login page

#auto download the excel file if the user computer dont have the excel file 
def generate_default_stock_list():
    """
    Generate a default stock list Excel file if it does not exist.
    """
    if os.path.exists(EXCEL_FILE_PATH):
        print("Default stock list already exists. Skipping generation.")
        return

    # Create a default DataFrame
    default_data = {
        "Category": ["Sample Category"],
        "Item Name": ["Sample Item"],
        "Supplier": ["Sample Supplier"],
        "Qty": [0],
        "Unit": ["pcs"],
        "Model": ["Sample Model"],
        "Packing Size": ["N/A"],
        "Date": ["--"]
    }
    df = pd.DataFrame(default_data)

    # Ensure the directory exists
    os.makedirs(os.path.dirname(EXCEL_FILE_PATH), exist_ok=True)

    # Save the DataFrame to the default path
    try:
        df.to_excel(EXCEL_FILE_PATH, index=False)
        push_to_github() 
        print("Default stock list generated.")
    except Exception as e:
        print(f"Error generating default stock list: {e}")

@app.route('/download-category/<sub_category_name>', methods=['GET'])
def download_category_excel(sub_category_name):
    try:
        sub_category_name = unquote(sub_category_name)

        if not os.path.exists(EXCEL_FILE_PATH):
            return "Stock data file not found.", 404

        # Load the Excel file and clean data
        df = pd.read_excel(EXCEL_FILE_PATH)
        df = clean_data(df)

        # Filter by sub-category
        filtered_df = df[df['Sub Category'].str.strip().str.lower() == sub_category_name.strip().lower()]

        # Ensure 'Stock Balance' is numeric
        filtered_df['Stock Balance'] = pd.to_numeric(filtered_df['Stock Balance'], errors='coerce').fillna(0)

        # Apply additional filters
        search_query = request.args.get('search', '').strip().lower()
        stock_status = request.args.get('stock_status', '').strip()
        supplier_filter = request.args.get('supplier', '').strip().lower()

        if search_query:
            filtered_df = filtered_df[filtered_df['Item Name'].str.lower().str.contains(search_query)]

        # Apply stock status filter
        if stock_status == 'Stock Low':
            filtered_df = filtered_df[filtered_df['Stock Balance'] <= 10]
        elif stock_status == 'Stock OK':
            filtered_df = filtered_df[filtered_df['Stock Balance'] > 10]

        # Apply supplier filter
        if supplier_filter:
            filtered_df = filtered_df[filtered_df['Supplier'].str.strip().str.lower() == supplier_filter]

        # Define columns to rename
        rename_columns = {
            'Sub Category': 'Category',
            'Item Name': 'Product Name',
            'Stock Balance': 'Quantity Available',
            # Add more renames here as needed
        }

        # Rename columns
        filtered_df = filtered_df.rename(columns=rename_columns)

        # Define columns to remove
        columns_to_exclude = [
            'Qty', 'Packing/Size', 'Min. Qty', 'Average Purchase Price',
            'Price Converted', 'China Cost (楼)', 'Local Item Link',
            'Picture', 'Last Updates', 'Stock Out'
            # Add more columns to remove here as needed
        ]

        # Drop unnecessary columns
        filtered_df = filtered_df.drop(columns=[col for col in columns_to_exclude if col in filtered_df.columns], errors='ignore')

        # Remove unnamed columns
        filtered_df = filtered_df.loc[:, ~filtered_df.columns.str.contains('^Unnamed')]

        # Save the filtered data to a temporary Excel file
        temp_dir = tempfile.mkdtemp()
        temp_file = os.path.join(temp_dir, f"{sanitize_filename(sub_category_name)}.xlsx")
        filtered_df.to_excel(temp_file, index=False)

        return send_file(temp_file, as_attachment=True, download_name=f"{sub_category_name}.xlsx")
    except Exception as e:
        return f"Error: {e}", 500

TECHNICAL_JSON_FILE = 'technical.json'
SITE_JSON_FILE = 'site.json'

@app.route('/inOutStock', methods=['GET', 'POST'])
def in_out_stock():
    if request.method == 'POST':
        # Retrieve form data
        item_name = request.form.get('item_name')
        quantity = int(request.form.get('quantity'))
        action = request.form.get('action')
        date = request.form.get('date') or datetime.now().strftime('%Y-%m-%d')  # Use current date if not provided
        who_taken = request.form.get('who_taken')
        site_name = request.form.get('site_name')
        remark = request.form.get('remark')

        # Load the Excel workbook
        if not os.path.exists(EXCEL_FILE_PATH):
            return "Stock List file not found.", 404

        workbook = load_workbook(EXCEL_FILE_PATH)
        push_to_github() 
        sheet1 = workbook["Sheet1"]
        sheet2 = workbook["Transaction"]
        sheet3 = workbook["updateSheet3"]  # Load Sheet3

        # Log the transaction in Sheet2
        sheet2_headers = [cell.value for cell in sheet2[1]]
        new_row_index_sheet2 = sheet2.max_row + 1
        transaction_log = {
        "Date": date,
        "Item Name": item_name,
        "Action": action,
        "Quantity": quantity,
        "Who Taken": who_taken,
        "Site Name": site_name,
        "Remark": remark
        }


        for header, value in transaction_log.items():
            if header in sheet2_headers:
                col_index = sheet2_headers.index(header) + 1
                sheet2.cell(row=new_row_index_sheet2, column=col_index, value=value)

        # Log the transaction in Sheet3
        sheet3_headers = [cell.value for cell in sheet3[1]]
        new_row_index_sheet3 = sheet3.max_row + 1
        for header, value in transaction_log.items():
            if header in sheet3_headers:
                col_index = sheet3_headers.index(header) + 1
                sheet3.cell(row=new_row_index_sheet3, column=col_index, value=value)

        # Update the stock balance in Sheet1
        sheet1_headers = [cell.value for cell in sheet1[1]]
        if "Item Name" in sheet1_headers and "Stock Balance" in sheet1_headers:
            item_name_col = sheet1_headers.index("Item Name") + 1
            qty_col = sheet1_headers.index("Stock Balance") + 1

            for row in range(2, sheet1.max_row + 1):
                if str(sheet1.cell(row=row, column=item_name_col).value).strip() == item_name.strip():
                    current_stock = sheet1.cell(row=row, column=qty_col).value or 0
                    if action == "In":
                        sheet1.cell(row=row, column=qty_col, value=current_stock + quantity)
                    elif action == "Out":
                        if current_stock >= quantity:
                            sheet1.cell(row=row, column=qty_col, value=current_stock - quantity)
                        else:
                            flash(f"Insufficient stock for '{item_name}'. Current stock: {current_stock}.", "error")
                            return render_template(
                                'inOutStock.html',
                                technicians=load_json_data(TECHNICAL_JSON_FILE),
                                sites=load_json_data(SITE_JSON_FILE)
                            )
                    break
            else:
                flash(f"Item '{item_name}' not found in Sheet1.", "error")
                return render_template(
                    'inOutStock.html',
                    technicians=load_json_data(TECHNICAL_JSON_FILE),
                    sites=load_json_data(SITE_JSON_FILE)
                )

        # Save changes to the workbook
        workbook.save(EXCEL_FILE_PATH)

        # Flash success message
        flash(f"Stock for '{item_name}' successfully updated!", "success")

        # Return success message
        return render_template(
            'inOutStock.html',
            technicians=load_json_data(TECHNICAL_JSON_FILE),
            sites=load_json_data(SITE_JSON_FILE)
        )

    # For GET requests
    technicians = load_json_data(TECHNICAL_JSON_FILE)
    sites = load_json_data(SITE_JSON_FILE)
    return render_template('inOutStock.html', technicians=technicians, sites=sites)

def load_json_data(file_path):
    if not os.path.exists(file_path):
        return []
    with open(file_path, 'r') as file:
        return json.load(file)

def save_json_data(file_path, data):
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

@app.route('/addTechnical.json', methods=['GET', 'POST', 'DELETE'])
def manage_technical():
    data = load_json_data(TECHNICAL_JSON_FILE)

    if request.method == 'POST':
        new_technical = request.json.get('name', '').strip()
        if new_technical and new_technical not in data:
            data.append(new_technical)
            save_json_data(TECHNICAL_JSON_FILE, data)
            push_to_github() 
            return jsonify({'status': 'success', 'message': 'Technical added successfully!', 'data': data})
        return jsonify({'status': 'error', 'message': 'Technical already exists or invalid input!'})

    if request.method == 'DELETE':
        name_to_delete = request.json.get('name', '').strip()
        if name_to_delete in data:
            data.remove(name_to_delete)
            save_json_data(TECHNICAL_JSON_FILE, data)
            return jsonify({'status': 'success', 'message': 'Technical deleted successfully!', 'data': data})
        return jsonify({'status': 'error', 'message': 'Technical not found!'})

    return jsonify(data)

@app.route('/addSite.json', methods=['GET', 'POST', 'DELETE'])
def manage_sites():
    data = load_json_data(SITE_JSON_FILE)

    if request.method == 'POST':
        new_site = request.json.get('name', '').strip()
        if new_site and new_site not in data:
            data.append(new_site)
            save_json_data(SITE_JSON_FILE, data)
            push_to_github() 
            return jsonify({'status': 'success', 'message': 'Site added successfully!', 'data': data})
        return jsonify({'status': 'error', 'message': 'Site already exists or invalid input!'})

    if request.method == 'DELETE':
        name_to_delete = request.json.get('name', '').strip()
        if name_to_delete in data:
            data.remove(name_to_delete)
            save_json_data(SITE_JSON_FILE, data)
            return jsonify({'status': 'success', 'message': 'Site deleted successfully!', 'data': data})
        return jsonify({'status': 'error', 'message': 'Site not found!'})

    return jsonify(data)

@app.route('/get_suggestions', methods=['GET'])
def get_suggestions():
    query = request.args.get('query', '').lower()
    if not query:
        return jsonify([])

    if os.path.exists(EXCEL_FILE_PATH):
        try:
            # Load the Excel workbook
            workbook = load_workbook(EXCEL_FILE_PATH, data_only=True)
            sheet1 = workbook["Sheet1"]

            # Extract only the "Item Name" column
            sheet1_data = sheet1.values
            sheet1_columns = next(sheet1_data)  # Extract header row
            item_name_col_index = sheet1_columns.index("Item Name")  # Find the column index of "Item Name"

            # Get all item names from the column
            item_names = [row[item_name_col_index] for row in sheet1.iter_rows(min_row=2, values_only=True)]

            # Remove duplicates and filter results based on the query
            unique_item_names = list(set(item_names))  # Remove duplicates
            filtered_item_names = [item for item in unique_item_names if item and query in item.lower()]

            return jsonify(filtered_item_names[:10])  # Return the top 10 results
        except Exception as e:
            print(f"Error fetching suggestions: {e}")
            return jsonify([])

    return jsonify([])

@app.route('/edit_excel', methods=['GET', 'POST'])
def edit_excel():
    if request.method == 'POST':
        # Get form data
        item_name = request.form.get('item_name').strip()
        min_qty = request.form.get('min_qty').strip()

        # Validate input
        if not item_name or not min_qty.isdigit():
            flash("Invalid input. Please provide a valid item name and minimum quantity.", "error")
            return redirect(url_for('edit_excel'))

        # Update the Excel file
        if os.path.exists(EXCEL_FILE_PATH):
            try:
                workbook = load_workbook(EXCEL_FILE_PATH)
                push_to_github() 
                sheet1 = workbook['Sheet1']

                # Locate the item and update the Min. Qty directly in the sheet
                headers = [cell.value for cell in sheet1[1]]  # Get headers from the first row
                if 'Item Name' in headers and 'Min. Qty' in headers:
                    item_col = headers.index('Item Name') + 1
                    min_qty_col = headers.index('Min. Qty') + 1

                    for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row):
                        if str(row[item_col - 1].value).strip().lower() == item_name.lower():
                            row[min_qty_col - 1].value = int(min_qty)
                            workbook.save(EXCEL_FILE_PATH)
                            flash(f"Updated 'Min. Qty' for '{item_name}' to {min_qty}.", "success")
                            return redirect(url_for('edit_excel'))

                    flash(f"Item '{item_name}' not found in Sheet1.", "error")
                else:
                    flash("Required columns 'Item Name' or 'Min. Qty' are missing in Sheet1.", "error")
            except Exception as e:
                flash(f"An error occurred while updating the stock: {str(e)}", "error")
        else:
            flash("Excel file not found.", "error")

        return redirect(url_for('edit_excel'))

    # For GET requests, render the HTML
    return render_template('edit_excel.html')



import os
from dotenv import load_dotenv
import subprocess
from threading import Lock
from subprocess import CalledProcessError, run

# Load environment variables
load_dotenv()

# GitHub configuration
GITHUB_USERNAME = "RSTechwin"
GITHUB_EMAIL = "rstechwinsetup@gmail.com"
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "").strip()  # Strip any whitespace or newline characters
GITHUB_REPO_URL = f"https://{GITHUB_TOKEN}@github.com/{GITHUB_USERNAME}/Stock-List-RS.git"

# Debugging: Print token information
if not GITHUB_TOKEN:
    print("Error: GitHub token is missing. Check your .env file.")
else:
    print(f"GitHub Token loaded: {GITHUB_TOKEN[:5]}... (truncated for security)")

git_lock = Lock()

def configure_git_user():
    """
    Configure Git username and email.
    """
    try:
        run(["git", "config", "--global", "user.name", GITHUB_USERNAME], check=True)
        run(["git", "config", "--global", "user.email", GITHUB_EMAIL], check=True)
    except CalledProcessError as e:
        print(f"Error configuring Git: {e}")

def remove_git_lock():
    """
    Remove Git lock file if it exists to prevent blocking Git operations.
    """
    lock_file = os.path.join(os.getcwd(), ".git", "index.lock")
    if os.path.exists(lock_file):
        os.remove(lock_file)
        print("Git lock file removed.")

def push_to_github():
    """
    Push updates to GitHub directly from the deployed application.
    """
    with git_lock:
        try:
            # Initialize Git repository if not already done
            if not os.path.exists(".git"):
                print("Initializing Git repository...")
                run(["git", "init"], check=True)

            # Configure Git user
            configure_git_user()

            # Set remote origin dynamically if not already set
            try:
                run(["git", "remote", "get-url", "origin"], check=True)
            except CalledProcessError:
                print("Setting remote origin...")
                run(["git", "remote", "add", "origin", GITHUB_REPO_URL], check=True)

            # Pull the latest changes to ensure no conflicts
            try:
                run(["git", "pull", "--rebase", "origin", "main"], check=True)
            except CalledProcessError as e:
                print(f"Warning: Could not pull changes: {e}")

            # Stage the changes
            print("Staging changes...")
            run(["git", "add", "."], check=True)

            # Commit the changes
            print("Committing changes...")
            run(["git", "commit", "-m", "Update stockList.xlsx"], check=True)

            # Push the changes
            print("Pushing changes to GitHub...")
            run(["git", "push", "-u", "origin", "main"], check=True)
            print("Changes pushed to GitHub successfully.")

        except CalledProcessError as e:
            print(f"Error pushing to GitHub: {e}")

if __name__ == '__main__':
    # Ensure GitHub token is loaded
    if not GITHUB_TOKEN:
        print("Error: GitHub token is missing. Check your .env file.")
    else:
        print(f"GitHub Token loaded: {GITHUB_TOKEN[:5]}... (truncated for security)")

    # Check if the Excel file exists
    if not os.path.exists(EXCEL_FILE_PATH):
        print("Excel file not found locally. Attempting to pull the latest version from GitHub...")

        try:
            # Clone the repository into a temporary directory
            temp_dir = tempfile.mkdtemp()
            subprocess.run(["git", "clone", GITHUB_REPO_URL, temp_dir], check=True)
            print("Repository cloned successfully.")

            # Move the Excel file to the desired location
            cloned_file_path = os.path.join(temp_dir, "files", "stockList.xlsx")
            if os.path.exists(cloned_file_path):
                os.makedirs(os.path.dirname(EXCEL_FILE_PATH), exist_ok=True)
                shutil.move(cloned_file_path, EXCEL_FILE_PATH)
                print(f"Excel file moved to: {EXCEL_FILE_PATH}")
            else:
                print("Error: stockList.xlsx not found in the cloned repository.")

        except subprocess.CalledProcessError as e:
            print(f"Error pulling stockList.xlsx from GitHub: {e}")
        except Exception as e:
            print(f"Unexpected error: {e}")
        finally:
            # Clean up the temporary directory
            shutil.rmtree(temp_dir, ignore_errors=True)
    else:
        print("Excel file found locally. Proceeding without pulling from GitHub.")

    # Start the Flask application
    app.run(host='0.0.0.0', port=5000, debug=True)


